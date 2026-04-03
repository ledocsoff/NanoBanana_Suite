"""
GeeLark XLSX Utilities — Shared helpers for reading/filling GeeLark templates.
"""

import os
import random
from datetime import datetime, timedelta, time as dtime
from typing import Any

# ─── Scheduling: Direct Paris Hours ────────────────────────────────
# Hours are PARIS time — written directly to XLSX, no conversion needed.

# Minimum delay between sequential tasks (warmup). Guard-rail.
MIN_SEQUENTIAL_DELAY = 15  # minutes

# Jitter safety factor: real usable capacity is ~70% of theoretical
# due to randomization/anti-pattern spreading.
_JITTER_SAFETY_FACTOR = 0.7


def range_duration_minutes(start_hour: int, end_hour: int) -> int:
    """Duration in minutes for a time range. Handles overnight natively.

    - start=20, end=4  → 480 min (8h overnight)
    - start=8,  end=20 → 720 min (12h)
    - start=0,  end=0  → 1440 min (24h full day)
    """
    if start_hour == end_hour:
        return 1440  # 24h
    if end_hour < start_hour:  # overnight
        return (24 - start_hour + end_hour) * 60
    return (end_hour - start_hour) * 60


def validate_schedule_capacity(
    total_tasks: int,
    start_hour: int,
    end_hour: int,
    min_gap_minutes: int,
    days_spread: int,
    max_simultaneous: int = 1,
) -> dict:
    """Validate if the schedule can fit all tasks. Returns diagnostics.

    Returns:
        {
            "feasible": bool,
            "window_minutes": int,
            "slots_per_day": int,
            "needed_days": int,
            "adjusted_days_spread": int,
            "warnings": list[str],
        }
    """
    window = range_duration_minutes(start_hour, end_hour)
    warnings = []

    # Alert 1: Very short window
    if window < 60:
        warnings.append(
            f"⚠️ Fenêtre très courte ({window} min). "
            f"Risque de clustering des posts."
        )

    # Effective capacity (apply jitter safety factor)
    raw_slots = max(1, window // min_gap_minutes) * max_simultaneous
    effective_slots = max(1, int(raw_slots * _JITTER_SAFETY_FACTOR))

    needed_days = -(-total_tasks // effective_slots)  # ceil division

    adjusted = days_spread
    if needed_days > days_spread:
        adjusted = needed_days
        warnings.append(
            f"🛠️ Capacité insuffisante : ~{effective_slots} slots effectifs/jour "
            f"pour {total_tasks} tâches. "
            f"Auto-ajustement days_spread : {days_spread} → {needed_days}."
        )

    # Alert 2: High density even after adjustment
    density = total_tasks / max(1, adjusted)
    if density > effective_slots * 0.8:
        warnings.append(
            f"⚠️ Densité élevée : ~{density:.0f} posts/jour "
            f"pour ~{effective_slots} slots effectifs. "
            f"Certains posts pourraient se chevaucher."
        )

    return {
        "feasible": needed_days <= 60,
        "window_minutes": window,
        "slots_per_day": effective_slots,
        "needed_days": needed_days,
        "adjusted_days_spread": adjusted,
        "warnings": warnings,
    }



def format_paris_time(dt: datetime) -> str:
    """Format a datetime as 'YYYY-MM-DD HH:MM' for GeeLark XLSX output.
    No timezone conversion — times are already in Paris hours."""
    return dt.strftime("%Y-%m-%d %H:%M")

# Column schemas for each GeeLark template type (1-indexed)
GEELARK_SCHEMAS = {
    "edit_profile": {
        "nickname": 5,
        "username": 6,
        "biography": 7,
        "link_url": 8,
        "link_title": 9,
    },
    "account_warmup": {
        "release_time": 4,
        "number_of_videos": 5,
        "search_keyword": 6,
    },
    "post_reel": {
        "caption": 5,
        "same_url": 6,
        "same_volume": 7,
        "acoustic_volume": 8,
        "ai_tags": 9,
    },
    "reels_gallery": {
        "caption": 5,
        "same_url": 6,
        "ai_tags": 7,
        "publish_post": 8,
    },
}

# Shared columns across all templates
COMMON_COLS = {
    "profile_serial": 1,
    "profile_name": 2,
    "task_no": 3,
    "release_time": 4,
}


def validate_template_type(ws, expected_type: str) -> bool:
    """Validate that the worksheet matches the expected GeeLark template type based on column headers."""
    headers = [str(cell.value).lower() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]
    col5 = headers[4] if len(headers) > 4 else ""
    
    actual_type = "post_video/carousel"
    if "nickname" in col5 or "username" in col5:
        actual_type = "edit_profile"
    elif "video" in col5 or "numberofvideo" in col5.replace(" ", ""):
        actual_type = "account_warmup"
        
    # Leniency for post templates since scheduler defaults to it
    if expected_type == "post_reel" or "post" in expected_type:
        if "post" not in actual_type:
            raise ValueError(f"❌ Erreur: Ce noeud attend un fichier 'Post Reel' mais a reçu '{actual_type}'.")
    elif actual_type != expected_type:
        raise ValueError(f"❌ Erreur: Ce noeud attend un fichier '{expected_type}' mais a reçu '{actual_type}'.\nVérifie que tu as sélectionné le bon export dans GeeLark !")
    return True


def load_template(path: str, expected_type: str = None):
    """Load a GeeLark XLSX template and return (workbook, data_rows)."""
    import openpyxl

    path = path.strip().strip("'\"")
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"Template introuvable: '{path}'")
    if not path.lower().endswith(".xlsx"):
        raise ValueError(f"Le template doit être un .xlsx: '{path}'")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    if expected_type:
        validate_template_type(ws, expected_type)
        
    data_rows = list(ws.iter_rows(min_row=2))
    return wb, data_rows


def fill_column(rows, col_index: int, values: list[str], randomize: bool = True):
    """Fill a specific column across all data rows with values from a pool.
    
    Args:
        rows: openpyxl row objects
        col_index: 1-indexed column number
        values: pool of values to assign
        randomize: shuffle the assignments
    """
    if not values:
        return

    if len(values) > 0 and len(rows) > 0:
        ratio = len(rows) / len(values)
        if ratio > 3:
            print(f"⚠️ [Attention] Seulement {len(values)} valeurs fournies pour {len(rows)} comptes (Duplication x{ratio:.1f})")

    pool = list(values)
    while len(pool) < len(rows):
        pool.extend(values)
    if randomize:
        random.shuffle(pool)
    pool = pool[:len(rows)]

    for row, val in zip(rows, pool):
        if val:
            row[col_index - 1].value = val


def fill_column_single(rows, col_index: int, value: str):
    """Fill a column with the same value for all rows."""
    if not value:
        return
    for row in rows:
        row[col_index - 1].value = value


def save_template(wb, output_path: str) -> str:
    """Save workbook and ensure the output directory exists."""
    output_path = output_path.strip().strip("'\"")
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def get_account_names(rows) -> list[str]:
    """Extract profile names from col 2."""
    return [str(row[1].value) if row[1].value else "unknown" for row in rows]
